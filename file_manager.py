from libraries import *
from spreadsheet import Spreadsheet
from cell import Cell
from fpdf import FPDF
from tabulate import tabulate

from spreadsheet_types import FileType, ExportType, ImportType
from spreadsheet_errors import *
from copy import deepcopy

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet



class FileManager:
    def __init__(self, spreadsheet: Spreadsheet) -> None:
        """
        This is class FileManager, It takes care of importing files to the spreadsheet and exporting files from a spreadsheet
        """
        self.spreadsheet = spreadsheet


    def export_to_file(self, file_name: str, file_type: FileType, mode: ExportType) -> None:
        """
        Function export_to_file
        The function handles the exporting of the current spreadsheet to an external file

        Parameters:
            * file_name: str
            * file_type: FileType - can be JSON, YAML, EXCEL, CSV, PDF.
            * mode: ExportType - can be FORMULAS, VALUES_ONLY

        Return Value: None
        """

        # Gettoing extension of the file - filetype
        splitted_name = file_name.split('.')
        extension = '.'+splitted_name[-1].strip()

        # Making sure filetype in the name is as wanted
        if extension != file_type.value:
            raise FileTypeError(f"Please Choose Valid extention, matched to your choice - {file_type.name}: {file_type.value}")

        # Dictionary of funftions based on file type given
        export_functions: Dict = {FileType.JSON: self.__export_to_json__, FileType.YAML: self.__export_to_yaml__, 
                                  FileType.CSV: self.__export_to_csv__, FileType.EXCEL: self.__export_to_excel__, FileType.PDF: self.__export_to_pdf__}

        # Validating file type 
        if file_type not in export_functions.keys():
            raise FileTypeError("File Type not supported")
    
        # Getting needed function
        export_function = export_functions[file_type]
    

        # Executing export using export function for each file
        if mode is ExportType.INCLUDE_INFO:
            
            # Making sure mode is supported for the wanted file
            if file_type is FileType.CSV or file_type is FileType.PDF:
                raise FileTypeError("File type can export values only, no cells info")

            # Export in order to reload - info export
            info_dict = {"rows": self.spreadsheet.rows, "cols": self.spreadsheet.cols, "cells": []}

            # Creating info dict
            for row in self.spreadsheet.cells_objs:
                for cell in row:
                    info = cell.get_info()
                    info_dict["cells"].append(deepcopy(info))

            # Exporting
            export_function(file_name, info=info_dict)

        elif mode is ExportType.VALUES_ONLY:
            # Exporting values only
            export_function(file_name)
        
        else:
            raise ExportTypeError("No such exporting exist")


    def import_file(self, file_name: str, file_type: FileType, mode: ExportType) -> None:
        """
        Function export_to_file
        The function handles the exporting of the current spreadsheet to an external file

        Parameters:
            * file_name: str
            * file_type: FileType - can be JSON, YAML, EXCEL, CSV, PDF.
            * mode: ExportType - can be FORMULAS, VALUES_ONLY

        Return Value: None
        """

        # Gettoing extension of the file - filetype
        splitted_name = file_name.split('.')
        extension = '.'+splitted_name[-1].strip()

        # Making sure filetype in the name is as wanted
        if extension != file_type.value:
            raise FileTypeError(f"Please Choose Valid extention, matched to your choice - {file_type.name}: {file_type.value}")

        # Dictionary of funftions based on file type given
        import_functions: Dict = {FileType.JSON: self.__import_from_json__, FileType.EXCEL: self.__import_from_excel__}
                                  

        # Validating file type 
        if file_type not in import_functions.keys():
            raise FileTypeError("File Type not supported")


        # Getting needed function
        import_func = import_functions[file_type]
        data = import_func(file_name, mode)

        if isinstance(data, Dict):
            try:
                # Extracting values
                cells = data['cells']
                rows = data['rows']
                cols = data['cols']

                self.spreadsheet.reset_sheet(rows, cols)

                for cell in cells:
                    index = cell['index']
                    curr_cell = self.spreadsheet.__get_cell_by_name__(cell['name'])

                    if curr_cell.index == index:
                        curr_cell.set_design(bg=cell['bg color'], fg=cell['fg color'])
                        self.spreadsheet.edit_cell(cell['name'], cell['formula'])
                    else:
                        self.spreadsheet.undo_reset()
                        raise FileFormatError("File is formatted incorrectly, index must be matched to cell's name. No file loaded.")

            except Exception as e:
                self.spreadsheet.undo_reset()
                raise FileFormatError("File Not formatted correctly, should be: {'rows': int, 'cols': int, 'cells': [Dicts of information per cell]}")
        elif isinstance(data, List):
            
            rows: int = data[0]
            cols: int = data[1]
            cells: Dict[str, Dict[str, str]] = data[2]
            
            self.spreadsheet.reset_sheet(rows, cols)
            
            for col in cells.keys():                
                for row, value in cells[col].items():                    
                    cell_id = col+row
                    new_val = value

                    try:
                        self.spreadsheet.edit_cell(cell_id, new_val)
                    except SpreadsheetError as error:
                        self.spreadsheet.undo_reset()
                        raise error


    def __export_to_json__(self, file_name: str, info={}) -> None:
        """
        Exports current spreadsheet to json file

        parameters:
            * file_name: str
            * info: dict, optional and just in case file being exported with info
        """

        if not info:
            # No formulas
            self.spreadsheet.cells_df.to_json(file_name)

        else:
            # Including formulas and values
            with open(file_name, 'w') as json_file:
                json.dump(info, json_file, indent=2)


    def __export_to_yaml__(self, file_name: str, info={}) -> None:
        """
        Exports current spreadsheet to yaml file

        parameters:
            * file_name: str
            * info: dict, optional and just in case file being exported with info
        """

        if not info:
            # Writing values only
            df_dict = self.spreadsheet.cells_df.to_dict()
            with open(file_name, 'w') as yaml_file:
                yaml.dump(df_dict, yaml_file, default_flow_style=False)
        
        else:
            # Writing all of the cell's info
            with open(file_name, 'w') as yaml_file:
                yaml.dump(info, yaml_file, default_flow_style=False)


    def __export_to_excel__(self, file_name: str, info={}) -> None:
        """
        Exports current spreadsheet to excel file

        parameters:
            * file_name: str
            * info: dict, optional and just in case file being exported with info
        """
        
        if info:
            # Breaking down dict values
            cells: List[Dict] = info["cells"]

            # Creating new excel file
            workbook = Workbook(file_name)
            worksheet = workbook.add_worksheet("sheet1")

            
            # Filling formulas in the right indexes
            for cell in cells:
                row = cell['index'][0]
                col = cell['index'][1]

                # Adding color formatting to the sheet
                cell_fg = cell['fg color']
                cell_bg = cell['bg color']


                format_with_color = workbook.add_format()
                format_with_color.set_font_color(cell_fg)

                if cell_bg != Cell.DEFULT_BG_COLOR:
                    format_with_color.set_bg_color(cell_bg)

                self.__value_to_excel__(worksheet, row, col, cell, format_with_color)

            # Closing file
            workbook.close()

        else:
            # Coping values only
            self.spreadsheet.cells_df.to_excel(file_name, index=False, header=False)


    def __value_to_excel__(self, worksheet: Workbook.worksheet_class, row, col, cell_info: Dict, cell_format) -> str:        
        """
        Function __value_to_excel__
        An auxiliary function that writes the cell's value to the excel sheet

        Parameters:
            * worksheet: Worksheet
            * row, col: int
            * cell_info: Dict
            * cell_format: Workbook.Format
        """

        # Making descision between formula and value and adding the cell's data to the right palce
        if self.spreadsheet.is_formula(cell_info['formula']):
            worksheet.write_formula(row, col - 1, cell_info["formula"], cell_format)
        else:
            value: str = cell_info['value']

            if value.isdigit():
                value = float(value) if '.' in value else int(value)

            worksheet.write(row, col - 1, value, cell_format)


    def __export_to_csv__(self, file_name: str) -> None:
        """
        Exports current spreadsheet to csv file - only by values, not by cell info

        parameters:
            * file_name: str
        """

        self.spreadsheet.cells_df.to_csv(file_name)


    def __export_to_pdf__(self, file_name: str) -> None:
        """
        Exports current spreadsheet to pdf file - only by values, no cell info

        parameters:
            * file_name: str
        """

        df = self.spreadsheet.cells_df

        # Convert DataFrame to a formatted table
        table = tabulate(df, headers='keys', tablefmt='grid')

        # Create PDF file
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', size=12)

        # Add table to PDF
        for line in table.split('\n'):
            pdf.cell(0, 10, txt=line, ln=True)

        # Save PDF file
        pdf.output(file_name)
    

    def __import_from_json__(self, filename: str, mode: ImportType) -> Union[Dict, List]:
        """
        Function __import_from_json__
        Imports from json file information based on the import type the user rquests.

        Parameters:
            * filename: str
            * mode - FORMULAS for full info or VALUES ONLY
        
        Return value: or a dict with all of the information, or a List of cells and their values
        """
    
        # Whole information including formulas and colors
        if mode == ImportType.FORMULAS:
            
            # Getting data while taking care of exceptions
            try:
                with open(filename, 'r') as json_file:
                    data: Dict = json.load(json_file)
            except:
                raise ImportFileError("File not Found")
            
            # Empty file, no info about size nor content
            if not data:
                raise EmptyFileImportError("Empty File, nothing imported. Loading empty sheet with defult values")
        
            return data

        # Getting only the values        
        elif mode == ImportType.VALUES_ONLY:
            try:
                with open(filename, 'r') as json_file:
                    data = json.load(json_file)

                # Getting neede values from data
                cols_num = len(data)

                cols_lens = []
                for col in data:
                    cols_lens.append(len(data[col]))

                # Getting the last row as the number of rows
                rows = max(cols_lens)

                return [rows, cols_num, data]

            except FileNotFoundError:
                raise ImportFileError("File not Found")
            except:
                raise FileFormatError('File was not formatted correctly')
        else:
            raise ImportTypeError('Import type used does not exist')


    def __import_from_excel__(self, filename: str, mode: ImportType) -> Dict:
        """
        Function __import_from_excel__
        Imports from excel file information, not including colors, but including all formulas

        Parameters:
            * filename: str
            * mode - FORMULAS for full info Only. cannot accept any other type
        
        Return value: or a dict with all of the information, or a List of cells and their values
        """

        if mode != ImportType.FORMULAS:
            raise ImportTypeError("Import Type error, only Formulas mode can use the function")

        try:
            workbook = openpyxl.load_workbook(filename)

        except FileNotFoundError:
            raise ImportFileError("File not found")
        

        # Getting info for dictionary
        worksheet: Worksheet = workbook.active
        total_rows = worksheet.max_row
        total_cols = worksheet.max_column

        # Creating base dict
        info_dict = {'rows': total_rows, 'cols': total_cols, 'cells': []}
        dict_fill = False

        # Getting cells information from the excel file
        for row in worksheet.iter_rows(values_only=False):
            
            for cell in row:
                
                # Getting values
                cell_value = cell.value if cell.value is not None else Cell.EMPTY_CELL
                cell_index = [cell.row-1, cell.column]
                cell_formula = cell_value # If it is a formula it already has the '=' prefix

                # Setting color to defult                
                cell_bg_color = Cell.DEFULT_BG_COLOR
                cell_color = Cell.DEFULT_FG_COLOR

                # Creating currnt dictionary
                cell_info = {'name': cell.column_letter+str(cell.row), 'index': cell_index, 'formula': cell_formula, 'value': cell_value, 'bg color': cell_bg_color, 'fg color': cell_color}
                info_dict['cells'].append(deepcopy(cell_info))

                # Covering case of empty excel file
                if not dict_fill:
                    dict_fill = True

        workbook.close()

        if not dict_fill:
            raise EmptyFileImportError("Empty File, Loading empty sheet with defult values")

        return info_dict
